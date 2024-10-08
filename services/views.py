import datetime
import mimetypes
import os
from time import strftime
from wsgiref.util import FileWrapper

from _ctypes import Structure
from django.contrib import auth
from django.utils import timezone
from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment, Font, NamedStyle
from pytils import translit
from django.contrib.auth.decorators import login_required
from django.shortcuts import render

from .models import *
from users.models import *
from django.contrib.auth.models import User

from django.http import HttpResponse, HttpResponseNotFound, StreamingHttpResponse, FileResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse_lazy
from django.views.generic import ListView, UpdateView, DetailView, DeleteView, CreateView

import pandas
# from pypdf import PdfFileReader, PdfReader
import locale
from docxtpl import DocxTemplate
from num2t4ru import decimal2text

# import aspose.words as aw

# Create your views here.

main_menu = [{'title': 'Главная', 'url_name': 'home', 'icon': 'fas fa-home fa-sm fa-fw me-2 text-gray-400',
              'role': ['Администраторы', 'Операторы', 'Экономисты',]},
             {'title': 'Создать отчет', 'url_name': 'dailyReport',
              'icon': 'fas fa-plus-square fa-sm fa-fw me-2 text-gray-400', 'role': ['Операторы']},
             {'title': 'Мои отчеты', 'url_name': 'myReports', 'icon': 'fas fa-file fa-sm fa-fw me-2 text-gray-400',
              'role': ['Операторы']},
             {'title': 'Мои услуги', 'url_name': 'myServices', 'icon': 'fas fa-tasks fa-sm fa-fw me-2 text-gray-400',
              'role': ['Операторы']},
             {'title': 'Отчеты', 'url_name': 'chooseStructure', 'icon': 'fas fa-file fa-sm fa-fw me-2 text-gray-400',
              'role': ['Администраторы', 'Экономисты']},
             ]

user_menu = [{'title': 'Профиль', 'url_name': 'profile', 'icon': 'fas fa-user fa-sm fa-fw me-2 text-gray-400',
              'role': ['Администраторы', 'Операторы', 'Экономисты', 'Библиотекари']},
             {'title': 'Админ-панель', 'url_name': 'admin:index', 'icon': 'fas fa-cogs fa-sm fa-fw me-2 text-gray-400',
              'role': ['Администраторы']},
             {'title': 'Выйти', 'url_name': 'logout', 'icon': 'fas fa-sign-out-alt fa-sm fa-fw me-2 text-gray-400',
              'role': ['Администраторы', 'Операторы', 'Экономисты', 'Библиотекари']}
             ]


def get_menu(groups):
    menu = []
    for group in groups:
        for menu_item in main_menu:
            if group in menu_item['role']:
                menu.append(menu_item)
    return menu


def get_user_menu(groups):
    menu = []
    for group in groups:
        for menu_item in user_menu:
            if group in menu_item['role']:
                menu.append(menu_item)
    return menu


@login_required
def home(request):
    # locale.setlocale(locale.LC_TIME, 'ru')
    user = auth.get_user(request)
    groups = request.user.groups.values_list('name', flat=True)
    context = {
        'menu': get_menu(groups),
        'user_menu': get_user_menu(groups),
        'title': 'ГрОИРО. Услуги',
        'userAvatar': Profile.objects.get(user=request.user).avatar,
        'groups': groups,
    }

    if 'Операторы' in groups:
        context['structure'] = Structures.objects.get(employee=user)
        reports = Reports.objects.filter(structure=context['structure'],
                                         date__gte=f"{timezone.now().year}-01-01").order_by('date')
        services = Services.objects.filter(
            structure=Structures.objects.get(employee=user))

        # График
        chart = {}
        for report in reports:
            if report.date.strftime("%B") not in chart:
                chart[report.date.strftime("%B")] = round(report.sum, 2)
            else:
                sum = chart[report.date.strftime("%B")]
                chart[report.date.strftime("%B")] = round(sum + report.sum, 2)
        context['chart'] = chart

        # Подсчет количества оказанных в текущем году услуг
        services_sum = {}
        for service in services:
            reports = Reports.objects.filter(service=service, date__gte=f"{timezone.now().year}-01-01")
            sum = 0
            for report in reports:
                sum += report.sum
            services_sum[service.name] = round(sum, 2)
        context['services_sum'] = services_sum

        if context['structure'].plan != 0:
            context['main_progress'] = round(context['structure'].earned / context['structure'].plan * 100, 1)
        else:
            context['main_progress'] = 0

    elif ('Экономисты' in groups or 'Администраторы' in groups):

        reports = Reports.objects.filter(date__gte=f"{timezone.now().year}-01-01").order_by('date')
        structures = Structures.objects.all()
        # Заработано средств
        sum = 0
        plan = 0
        for structure in structures:
            sum += structure.earned
            plan += structure.plan
        context['earned'] = round(sum, 2)
        context['plan'] = round(plan, 2)

        # Выполнение плана
        if plan != 0:
            context['main_progress'] = round(sum / plan * 100, 1)
        else:
            context['main_progress'] = 0

        # График
        chart = {}
        for report in reports:
            if report.date.strftime("%B") not in chart:
                chart[report.date.strftime("%B")] = round(report.sum, 2)
            else:
                sum = chart[report.date.strftime("%B")]
                chart[report.date.strftime("%B")] = round(sum + report.sum, 2)
        context['chart'] = chart

        # Выполнение по каждой услуге
        services_sum = {}
        for service in Services.objects.all():
            reports = Reports.objects.filter(service=service, date__gte=f"{timezone.now().year}-01-01")
            sum = 0
            for report in reports:
                sum += report.sum
            services_sum[service.name] = round(sum, 2)
        context['services_sum'] = services_sum

    # if request.user.is_authenticated:
    #     context['userRole'] = group

    if request.method == 'POST':
        operator_name = str(user.first_name)
        operator_surname = str(user.last_name)
        IOF_operator = operator_name.split(' ')[0][0] + '.' + operator_name.split(' ')[1][0] + '.' + operator_surname
        FIO_operator = operator_surname + ' ' + operator_name.split(' ')[0][0] + '.' + operator_name.split(' ')[1][
            0] + '.'

        date = request.POST.get('date').split('-')[2] + '.' + request.POST.get('date').split('-')[1] + '.' + \
               request.POST.get('date').split('-')[0]

        reports_to_print = Reports.objects.filter(structure=Structures.objects.get(employee=user),
                                                  date=request.POST.get('date'))
        int_units = ((u'рубль', u'рубля', u'рублей'), 'm')
        exp_units = ((u'копейка', u'копейки', u'копеек'), 'f')
        sum_cash = 0
        nds_cash = 0
        sum_card = 0
        nds_card = 0
        servs_cash = ''
        servs_card = ''
        for i, r in enumerate(reports_to_print):
            if r.cash_sum > 0:
                sum_cash += r.cash_sum
                nds_cash += r.cash_nds
                servs_cash += f'\t{i + 1}. {r.service.name} {(str(r.cash_sum).split(".")[0])} руб. {str("{:.2f}".format(r.cash_sum)).split(".")[1]} коп. ({decimal2text(r.cash_sum, int_units=int_units, exp_units=exp_units)}), ' \
                              f'в т.ч. НДС - {str(r.cash_nds).split(".")[0]} руб. {str("{:.2f}".format(r.cash_nds)).split(".")[1]} коп. ({decimal2text(r.cash_nds, int_units=int_units, exp_units=exp_units)}).\n'
            if r.card_sum >0:
                sum_card += r.card_sum
                nds_card += r.card_nds
                servs_card += f'\t{i + 1}. {r.service.name} {(str(r.card_sum).split(".")[0])} руб. {str("{:.2f}".format(r.card_sum)).split(".")[1]} коп. ({decimal2text(r.card_sum, int_units=int_units, exp_units=exp_units)}), ' \
                              f'в т.ч. НДС - {str(r.card_nds).split(".")[0]} руб. {str("{:.2f}".format(r.card_nds)).split(".")[1]} коп. ({decimal2text(r.card_nds, int_units=int_units, exp_units=exp_units)}).\n'

        data = {
            'res_cash': f'Всего: {str("{:.2f}".format(sum_cash)).split(".")[0]} руб. {str("{:.2f}".format(sum_cash)).split(".")[1]} коп. ({decimal2text(sum_cash, int_units=int_units, exp_units=exp_units)}), '
                   f'в т.ч. НДС – {str("{:.2f}".format(nds_cash)).split(".")[0]} руб. {str("{:.2f}".format(nds_cash)).split(".")[1]} коп.  ({decimal2text(nds_cash, int_units=int_units, exp_units=exp_units)}).',
            'res_card': f'Всего: {str("{:.2f}".format(sum_card)).split(".")[0]} руб. {str("{:.2f}".format(sum_card)).split(".")[1]} коп. ({decimal2text(sum_card, int_units=int_units, exp_units=exp_units)}), '
                        f'в т.ч. НДС – {str("{:.2f}".format(nds_card)).split(".")[0]} руб. {str("{:.2f}".format(nds_card)).split(".")[1]} коп.  ({decimal2text(nds_card, int_units=int_units, exp_units=exp_units)}).',
            'date': date,
            'IOF_operator': IOF_operator,
            'FIO_operator': FIO_operator,
            'servs_cash': servs_cash,
            'servs_card': servs_card,
        }
        # if sum_cash > 0:
        #     data['res_cash'] = (f'Всего: {str("{:.2f}".format(sum_cash)).split(".")[0]} руб. '
        #                         f'{str("{:.2f}".format(sum_cash)).split(".")[1]} коп. '
        #                         f'({decimal2text(sum_cash, int_units=int_units, exp_units=exp_units)}), '
        #                         f'в т.ч. НДС – {str("{:.2f}".format(nds_cash)).split(".")[0]} руб. '
        #                         f'{str("{:.2f}".format(nds_cash)).split(".")[1]} коп.  '
        #                         f'({decimal2text(nds_cash, int_units=int_units, exp_units=exp_units)}).')
        #
        # if sum_card > 0:
        #     data['res_card'] = (f'Всего: {str("{:.2f}".format(sum_card)).split(".")[0]} руб. '
        #                         f'{str("{:.2f}".format(sum_card)).split(".")[1]} коп. '
        #                         f'({decimal2text(sum_card, int_units=int_units, exp_units=exp_units)}), '
        #                         f'в т.ч. НДС – {str("{:.2f}".format(nds_card)).split(".")[0]} руб. '
        #                         f'{str("{:.2f}".format(nds_card)).split(".")[1]} коп.  '
        #                         f'({decimal2text(nds_card, int_units=int_units, exp_units=exp_units)}).')

        doc = DocxTemplate("services/static/template.docx")

        filename = f"Отчет {data['date']}.docx"
        filepath = f'services/reports/{filename}'
        doc.render(data)
        doc.save(filepath)

        chunk_size = 8192

        response = StreamingHttpResponse(FileResponse(open(filepath, 'rb'), chunk_size),
                                         content_type=mimetypes.guess_type(filepath)[0])

        response['Content-Length'] = os.path.getsize(filepath)
        response['Content-Disposition'] = "attachment; filename=%s" % filename

        return response
        # return redirect('home')


    return render(request, 'index.html', context)


def update_earned(request):
    sum = 0
    user = auth.get_user(request)
    groups = request.user.groups.values_list('name', flat=True)
    reports = Reports.objects.filter(structure=Structures.objects.get(employee=user),
                                     date__gte=f"{timezone.now().year}-01-01")
    for report in reports:
        sum += report.sum
    updSum = Structures.objects.get(employee=user)
    updSum.earned = sum
    updSum.save(update_fields=['earned'])
    return redirect(home)


@login_required
def dailyReport(request):
    user = auth.get_user(request)
    groups = request.user.groups.values_list('name', flat=True)
    context = {
        'menu': get_menu(groups),
        'user_menu': get_user_menu(groups),
        'title': 'Ежедневный отчет',
        'services': Services.objects.filter(
            structure=Structures.objects.get(employee=user)),
        'structure': Structures.objects.filter(employee=user)[0],
        'current_date': timezone.now(),
        'userAvatar': Profile.objects.get(user=request.user).avatar,
    }
    if request.method == 'POST':
        date = request.POST.get('service_date')
        service_id_list = request.POST.getlist('service_id')
        service_count_cash_list = request.POST.getlist('service_count_cash')
        service_count_card_list = request.POST.getlist('service_count_card')

        for i in range(len(service_id_list)):
            cash_sum = round(int(service_count_cash_list[i]) * Services.objects.filter(pk=service_id_list[i])[0].price,
                             2)
            card_sum = round(int(service_count_card_list[i]) * Services.objects.filter(pk=service_id_list[i])[0].price,
                             2)

            if int(service_count_cash_list[i]) != 0 or int(service_count_card_list[i]) != 0:
                try:
                    report = Reports.objects.get(date=date, service=Services.objects.get(pk=service_id_list[i]),
                                                 structure=Structures.objects.filter(employee=user)[0])

                    cash_amount = report.cash_amount + int(service_count_cash_list[i])
                    report.cash_amount = cash_amount
                    cash_sum = report.cash_sum + cash_sum
                    report.cash_sum = cash_sum
                    cash_nds = report.cash_nds + round(cash_sum / 6, 2)
                    report.cash_nds = cash_nds

                    card_amount = report.card_amount + int(service_count_card_list[i])
                    report.card_amount = card_amount
                    card_sum = report.card_sum + card_sum
                    report.card_sum = card_sum
                    card_nds = report.card_nds + round(card_sum / 6, 2)
                    report.card_nds = card_nds

                    report.amount = cash_amount + card_amount
                    report.sum = cash_sum + card_sum
                    report.nds = cash_nds + card_nds

                    report.save(
                        update_fields=['cash_amount', 'cash_sum', 'cash_nds', 'card_amount',
                                       'card_sum', 'card_nds', 'amount', 'sum', 'nds'])

                except Reports.DoesNotExist:
                    Reports.objects.create(
                        date=date,
                        service=Services.objects.get(pk=service_id_list[i]),
                        structure=Structures.objects.filter(employee=user)[0],

                        cash_amount=service_count_cash_list[i],
                        cash_sum=cash_sum,
                        cash_nds=round(cash_sum / 6, 2),

                        card_amount=service_count_card_list[i],
                        card_sum=card_sum,
                        card_nds=round(card_sum / 6, 2),

                        amount=int(service_count_cash_list[i]) + int(service_count_card_list[i]),
                        sum=cash_sum + card_sum,
                        nds=round((cash_sum + card_sum) / 6, 2)
                    )

                updSum = Structures.objects.get(employee=user)
                updSum.earned = updSum.earned + cash_sum + card_sum
                updSum.save(update_fields=['earned'])
            else:
                continue

        return redirect('home')

    return render(request, 'DailyReport.html', context)


@login_required
def chooseStructure(request):
    user = auth.get_user(request)
    groups = request.user.groups.values_list('name', flat=True)
    context = {
        'menu': get_menu(groups),
        'user_menu': get_user_menu(groups),
        'title': 'Выбор структуры',
        'structures': Structures.objects.all(),
        'userAvatar': Profile.objects.get(user=request.user).avatar,
    }

    return render(request, 'chooseStructure.html', context)


@login_required
def listReports(request, pk):
    user = auth.get_user(request)
    groups = request.user.groups.values_list('name', flat=True)
    context = {
        'menu': get_menu(groups),
        'user_menu': get_user_menu(groups),
        'title': 'Отчеты',
        'structure': Structures.objects.filter(pk=pk).first,
        'services': Services.objects.filter(structure=Structures.objects.get(pk=pk)),
        'reports': Reports.objects.filter(structure=pk, date__gte=f"{timezone.now().year}-01-01").order_by('-date'),
        'userAvatar': Profile.objects.get(user=request.user).avatar,
    }

    if request.method == 'POST':
        date_start = str(request.POST.get('date_start'))
        date_finish = str(request.POST.get('date_finish'))
        service = request.POST.get('service')
        print(f'c _{date_start}_ po _{date_finish}_')

        if date_start and date_finish != '' and service != 'all':
            print(date_start, date_finish, service)
            context['reports'] = Reports.objects.filter(structure=pk, service=Services.objects.get(pk=service),
                                                        date__gte=date_start,
                                                        date__lte=date_finish).order_by('-date')
            context['date_start'] = date_start
            context['date_finish'] = date_finish
            context['service'] = Services.objects.get(pk=service)


        elif date_start != '' and date_finish != '' and service == 'all':
            print(date_start, date_finish, service)
            context['reports'] = Reports.objects.filter(structure=pk, date__gte=date_start,
                                                        date__lte=date_finish).order_by('-date')
            context['date_start'] = date_start
            context['date_finish'] = date_finish

        elif date_start == '' and date_finish == '' and service != 'all':
            print(service)
            context['reports'] = Reports.objects.filter(structure=pk, service=service).order_by('-date')
            context['service'] = Services.objects.get(pk=service)

        else:
            print('date_start', 'date_finish', service)
            context['reports'] = Reports.objects.filter(structure=pk).order_by('-date')

    return render(request, 'listReports.html', context)


def downloadReport(request):
    user = auth.get_user(request)
    groups = request.user.groups.values_list('name', flat=True)
    service = request.GET.get('service')
    date_start = request.GET.get('date_start')
    date_finish = request.GET.get('date_finish')
    structure = request.GET.get('structure')

    if 'Операторы' in groups:
        pk = Structures.objects.get(employee=user).pk
        reports = Reports.objects.filter(structure=Structures.objects.get(pk=pk))
    elif ('Экономисты' in groups or
          'Администраторы' in groups):
        reports = Reports.objects.filter(structure=Structures.objects.get(name=structure))

    if date_start and date_finish != '' and service != 'all':

        reports = reports.filter(service=Services.objects.get(pk=service),
                                 date__gte=date_start,
                                 date__lte=date_finish)

    elif date_start and date_finish != '' and service == 'all':

        reports = reports.filter(
            date__gte=date_start,
            date__lte=date_finish)

    elif date_start and date_finish == '' and service != 'all':

        reports = reports.filter(
            service=Services.objects.get(pk=service))

    else:
        reports = reports

    dates = []
    services = []
    cash_amounts = []
    cash_sums = []
    card_amounts = []
    card_sums = []
    total_amounts = []
    total_sums = []

    for r in reports:
        dates.append(str(r.date))
        services.append(str(r.service))
        cash_amounts.append(int(r.cash_amount))
        cash_sums.append(float(r.cash_sum))
        card_amounts.append(int(r.card_amount))
        card_sums.append(float(r.card_sum))
        total_amounts.append(int(r.amount))
        total_sums.append(float(r.sum))

    dates.append('')
    cash_amounts.append(f'=SUM(C3:C{len(cash_amounts) + 2})')
    cash_sums.append(f'=SUM(D3:D{len(cash_sums) + 2})')
    card_amounts.append(f'=SUM(E3:E{len(card_amounts) + 2})')
    card_sums.append(f'=SUM(F3:F{len(card_sums) + 2})')
    total_amounts.append(f'=SUM(G3:G{len(total_amounts) + 2})')
    total_sums.append(f'=SUM(H3:H{len(total_sums) + 2})')
    services.append('Итого:')

    # Вариант pandas
    # data = {
    #     ('', 'Дата'): dates,
    #     ('', 'Услуга'): services,
    #     ('Наличный рассчет', 'Количество'): cash_amounts,
    #     ('Наличный рассчет', 'Сумма'): cash_sums,
    #     ('Безналичный рассчет', 'Количество'): card_amounts,
    #     ('Безналичный рассчет', 'Сумма'): cash_sums,
    #     ('Итого', 'Количество'): total_amounts,
    #     ('Итого', 'Сумма'): total_sums,
    # }
    # df = pandas.DataFrame(data)

    # df = pandas.DataFrame({'Дата': dates,
    #                        'Услуга': services,
    #                        'Нал. количество': cash_amounts,
    #                        'Нал. сумма': cash_sums,
    #                        'Безнал. количество': card_amounts,
    #                        'Безнал. сумма': card_sums,
    #                        'Общее количество': total_amounts,
    #                        'Общая сумма': total_sums,
    #                        })
    if date_start and date_finish == '':
        filename = f'Отчет_all_time.xlsx'
    else:
        filename = f'Отчет_{date_start}-{date_finish}.xlsx'

    filepath = f'services/reports/{filename}'
    # df.to_excel(filepath, index=True)

    # Вариант openpyxl
    wb = Workbook()
    ws = wb.active

    ws['A1'] = 'Дата'
    ws['B1'] = 'Услуга'
    ws['C1'] = 'Наличный расчет'
    ws['C2'] = 'Количество'
    ws['D2'] = 'Сумма'
    ws['E1'] = 'Безналичный расчет'
    ws['E2'] = 'Количество'
    ws['F2'] = 'Сумма'
    ws['G1'] = 'Итого'
    ws['G2'] = 'Количество'
    ws['H2'] = 'Сумма'

    ws.merge_cells('C1:D1')
    ws.merge_cells('E1:F1')
    ws.merge_cells('G1:H1')
    ws.merge_cells('A1:A2')
    ws.merge_cells('B1:B2')

    data2 = {
        'A': dates,
        'B': services,
        'C': cash_amounts,
        'D': cash_sums,
        'E': card_amounts,
        'F': card_sums,
        'G': total_amounts,
        'H': total_sums,
    }

    # date_style = NamedStyle(name='date_style', number_format='YYYY.MM.DD')
    # wb.add_named_style(date_style)

    for col, data in data2.items():
        for i, value in enumerate(data, start=3):
            cell = ws[f'{col}{i}']
            cell.value = value
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if i == len(data) + 2 and col != 'A':
                cell.font = Font(bold=True)
            # if col == 'A' and value is not None:
            #     cell.style = date_style

    for cell in ['A1', 'B1', 'C1', 'E1', 'G1', 'C2', 'D2', 'E2', 'F2', 'G2', 'H2']:
        ws[cell].alignment = Alignment(horizontal='center', vertical='center')
        ws[cell].font = Font(bold=True)

    wb.save(filepath)

    chunk_size = 8192

    response = StreamingHttpResponse(FileResponse(open(filepath, 'rb'), chunk_size),
                                     content_type=mimetypes.guess_type(filepath)[0])

    response['Content-Length'] = os.path.getsize(filepath)
    response['Content-Disposition'] = "attachment; filename=%s" % filename

    return response


@login_required
def myReports(request):
    user = auth.get_user(request)
    groups = request.user.groups.values_list('name', flat=True)
    context = {
        'menu': get_menu(groups),
        'user_menu': get_user_menu(groups),
        'title': 'Мои отчеты',
        'services': Services.objects.filter(
            structure=Structures.objects.get(employee=user)),
        'structure': Structures.objects.filter(employee=user).first,
        'reports': Reports.objects.filter(
            structure=Structures.objects.get(employee=user),
            date__gte=f"{timezone.now().year}-01-01").order_by('-date'),
        'userAvatar': Profile.objects.get(user=request.user).avatar,
    }
    if request.method == 'POST':
        pk = Structures.objects.get(employee=user)
        date_start = str(request.POST.get('date_start'))
        date_finish = str(request.POST.get('date_finish'))
        service = request.POST.get('service')

        if date_start and date_finish != '' and service != 'all':
            context['reports'] = Reports.objects.filter(structure=pk, service=Services.objects.get(pk=service),
                                                        date__gte=date_start,
                                                        date__lte=date_finish).order_by('-date')
            context['date_start'] = date_start
            context['date_finish'] = date_finish
            context['service'] = Services.objects.get(pk=service)


        elif date_start != '' and date_finish != '' and service == 'all':
            print(date_start, date_finish, service)
            context['reports'] = Reports.objects.filter(structure=pk, date__gte=date_start,
                                                        date__lte=date_finish).order_by('-date')
            context['date_start'] = date_start
            context['date_finish'] = date_finish

        elif date_start == '' and date_finish == '' and service != 'all':
            print(service)
            context['reports'] = Reports.objects.filter(structure=pk, service=service).order_by('-date')
            context['service'] = Services.objects.get(pk=service)

        else:
            print('date_start', 'date_finish', service)
            context['reports'] = Reports.objects.filter(structure=pk).order_by('-date')

    return render(request, 'listReports.html', context)


@login_required
def myServices(request):
    user = auth.get_user(request)
    groups = request.user.groups.values_list('name', flat=True)
    context = {
        'menu': get_menu(groups),
        'user_menu': get_user_menu(groups),
        'title': 'Мои услуги',
        'services': Services.objects.filter(
            structure=Structures.objects.get(employee=user)),
        # 'structure': Structures.objects.filter(employee=Profile.objects.get(user=request.user)).first,
        'userAvatar': Profile.objects.get(user=request.user).avatar,

    }

    if request.method == 'POST':
        pks = request.POST.getlist('pk')
        names = request.POST.getlist('name')
        descriptions = request.POST.getlist('description')
        prices = request.POST.getlist('price')

        for i in range(len(names)):
            Services.objects.filter(pk=pks[i]).update(name=names[i], description=descriptions[i], price=prices[i])

        return redirect('myServices')

    return render(request, 'my_services.html', context)


def del_service(request, pk):
    Services.objects.get(pk=pk).delete()
    return redirect(myServices)


@login_required
def newService(request):
    user = auth.get_user(request)
    groups = request.user.groups.values_list('name', flat=True)
    context = {
        'menu': get_menu(groups),
        'user_menu': get_user_menu(groups),
        'title': 'Новая услуга',
        'structure': Structures.objects.filter(employee=user).first,
    }
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description')
        price = request.POST.get('price')

        Services.objects.create(name=name, description=description, price=price,
                                structure=Structures.objects.filter(employee=user)[0])
        return redirect('myServices')

    return render(request, 'newService.html', context)

# @login_required
# def ibcNewOrder(request):
#     price = Services.objects.get(name='Печать').price
#     context = {
#         'title': 'Новый заказ',
#         'price': price
#     }
#
#     if request.method == 'POST':
#         name = request.POST.get('name')
#         email = request.POST.get('email')
#         phone = request.POST.get('phone')
#         date_of_receiving = request.POST.get('date_of_receiving')
#         document = str(request.FILES['document']).replace(' ', '_')
#         comment = request.POST.get('comment')
#         #
#         # count_pages = len(PdfReader(document).pages)
#         # print(f'{count_pages} страниц')
#
#         order = Orders(service=Services.objects.get(pk=1),
#                        client_name=name,
#                        sum=0,
#                        email=email,
#                        phone=phone,
#                        date_of_receiving=date_of_receiving,
#                        comment=comment,
#                        file=request.FILES['document'],
#                        status='Заказ принят')
#         order.save()
#         print('save ok ----> ', document)
#         # doc = aw.Document(f'./uploads/ordersIBC/{strftime("%Y")}/{strftime("%m")}/{strftime("%d")}/{document}')
#         # Orders.objects.filter(pk=order.id).update(sum=price * doc.page_count)
#         # print('update ok ---> ', document)
#
#     return render(request, 'ibcNewOrder.html', context)
